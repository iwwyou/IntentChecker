"""
Assemble the RQ2-A / RQ1 specification-profile Excel workbook from:
  - evaluation/RQ2/extracted/<case_id>.json   (per-case profile, 4 groups + excluded)
  - evaluation/RQ2/rq2b_latency.csv           (10-run latency, 45 cases)
  - evaluation/RQ2/rq2b_cumulative_median.csv (stability check, 45 cases)
  - evaluation/RQ2/rq2b_validation.json       (RQ1-B validation outcome, 45 cases)

Output: evaluation/RQ2/rq2a_specification_profile.xlsx
  Sheets: raw_profile, latency, aggregates, by_outcome, value_usable_matrix

Usage:
    .venv/Scripts/python.exe evaluation/RQ2/build_excel.py
"""

import csv
import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent.parent
RQ2_DIR = PROJECT_ROOT / "evaluation" / "RQ2"
EXTRACTED_DIR = RQ2_DIR / "extracted"
LATENCY_CSV = RQ2_DIR / "rq2b_latency.csv"
CUMMEDIAN_CSV = RQ2_DIR / "rq2b_cumulative_median.csv"
VALIDATION_JSON = RQ2_DIR / "rq2b_validation.json"
STRUCTURAL_CSV = RQ2_DIR / "rq1_structural.csv"
OUT_XLSX = RQ2_DIR / "rq2a_specification_profile.xlsx"

# The one case known to be excluded from the eligible 75->74 pool.
EXCLUDED_CASE_ID = "web3bugs_101_H_02"

RAW_COLUMNS = [
    "case_id", "group", "case_built", "expressible", "value_or_algorithm", "usable",
    "blocker_tags", "relation_form", "during_or_post", "context_breadth",
    "relevant_statements", "unique_values", "unique_values_breakdown",
    "additional_functions_n", "additional_functions_note",
    "additional_contracts_n", "additional_contracts_note",
    "external_spec", "external_spec_note", "annotation_multiplicity",
    "finding_level_note", "case_notes", "annotation_text",
    "validation_outcome", "validation_note",
    "latency_median_s", "latency_iqr_low_s", "latency_iqr_high_s",
    "final_class", "triage_status", "reclass_reason", "notes",
    "source_file", "extraction_note",
]


def load_extracted():
    records = {}
    if not EXTRACTED_DIR.exists():
        return records
    for f in sorted(EXTRACTED_DIR.glob("*.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[warn] failed to parse {f}: {e}")
            continue
        cid = data.get("case_id", f.stem)
        records[cid] = data
    return records


def load_latency():
    out = {}
    if not LATENCY_CSV.exists():
        return out
    with open(LATENCY_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["case_id"]] = row
    return out


def load_cummedian():
    out = {}
    if not CUMMEDIAN_CSV.exists():
        return out
    with open(CUMMEDIAN_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["case_id"]] = row
    return out


def load_validation():
    if not VALIDATION_JSON.exists():
        return {}
    return json.loads(VALIDATION_JSON.read_text(encoding="utf-8"))


def merge_records(extracted, latency, validation):
    """Return list of row-dicts, one per known case_id, unified across groups."""
    rows = []
    all_ids = set(extracted.keys())
    for cid, rec in sorted(extracted.items()):
        row = {k: rec.get(k, "") for k in RAW_COLUMNS}
        row["case_id"] = cid

        lat = latency.get(cid)
        if lat:
            row["latency_median_s"] = lat.get("median", "")
            row["latency_iqr_low_s"] = lat.get("q1", "")
            row["latency_iqr_high_s"] = lat.get("q3", "")

        val = validation.get(cid)
        if val:
            # Prefer a fresh engine-observed outcome unless the extracted record
            # already carries a hand-curated validation_outcome (e.g. numscout_EthereumGod's
            # multi-member A: Violated; B: Unsupported note from the analysis.md).
            if not row.get("validation_outcome"):
                row["validation_outcome"] = val.get("case_level_outcome", "")
            if not row.get("validation_note") and val.get("missing_lines"):
                row["validation_note"] = (
                    f"{val['missing_lines']} of {val['declared_annotations']} declared "
                    f"annotation(s) produced no output line (silent/Unsupported)"
                )
        rows.append(row)

    return rows


def autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 60)


def write_raw_profile(wb, rows):
    ws = wb.active
    ws.title = "raw_profile"
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    ws.append(RAW_COLUMNS)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
    for row in rows:
        ws.append([row.get(k, "") for k in RAW_COLUMNS])
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def to_int(v):
    try:
        iv = int(v)
        return iv if iv >= 0 else None
    except (TypeError, ValueError):
        return None


def quartiles(sorted_vals):
    n = len(sorted_vals)
    if n == 0:
        return None, None
    def median_of(vals):
        m = len(vals)
        if m == 0:
            return None
        mid = m // 2
        if m % 2 == 1:
            return vals[mid]
        return (vals[mid - 1] + vals[mid]) / 2
    mid = n // 2
    lower = sorted_vals[:mid]
    upper = sorted_vals[mid + (n % 2):]
    return median_of(lower), median_of(upper)


def write_aggregates(wb, rows):
    ws = wb.create_sheet("aggregates")
    bold = Font(bold=True)

    # RQ2-A eligible set = every Expressible=Yes case (baseline20 + ALL phase_reviews_expressible,
    # including any not yet case-built -- RQ2-A applies to Expressible cases regardless of build
    # status, README §6). This is 46 cases (20 + 26), NOT the same as the 45-case validated set.
    rq2a_rows = [r for r in rows if r["group"] in ("baseline20", "phase_reviews_expressible")]
    # RQ1-B/RQ2-B validated set = only cases with an actual executed case JSON (45 cases: 20
    # baseline + 25 case-built phase_reviews_expressible). Use this for anything involving
    # validation_outcome/latency -- a case with no run has no outcome to cross-tabulate.
    validated_rows = [r for r in rq2a_rows if r.get("case_built") == "Yes"]
    n_unbuilt = len(rq2a_rows) - len(validated_rows)
    if n_unbuilt:
        ws_note_unbuilt = n_unbuilt
    else:
        ws_note_unbuilt = 0

    def stats_block(title, key, r=1):
        vals = sorted(v for v in (to_int(row.get(key)) for row in rq2a_rows) if v is not None)
        ws.cell(r, 1, title).font = bold
        if vals:
            q1, q3 = quartiles(vals)
            ws.cell(r, 2, "median"); ws.cell(r, 3, statistics.median(vals))
            ws.cell(r+1, 2, "Q1"); ws.cell(r+1, 3, q1)
            ws.cell(r+2, 2, "Q3"); ws.cell(r+2, 3, q3)
            ws.cell(r+3, 2, "min"); ws.cell(r+3, 3, min(vals))
            ws.cell(r+4, 2, "max"); ws.cell(r+4, 3, max(vals))
            ws.cell(r+5, 2, "n"); ws.cell(r+5, 3, len(vals))
        return r + 7

    row_ptr = 1
    row_ptr = stats_block("Relevant statements (median/IQR) -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)", "relevant_statements", row_ptr)
    row_ptr = stats_block("Unique relevant values (median/IQR) -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)", "unique_values", row_ptr)

    # Context breadth histogram + locality buckets
    ws.cell(row_ptr, 1, "Context breadth histogram (0-4) -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)").font = bold
    row_ptr += 1
    breadth_counts = {i: 0 for i in range(5)}
    for row in rq2a_rows:
        b = to_int(row.get("context_breadth"))
        if b is not None and 0 <= b <= 4:
            breadth_counts[b] += 1
    for b in range(5):
        ws.cell(row_ptr, 2, f"breadth={b}"); ws.cell(row_ptr, 3, breadth_counts[b])
        row_ptr += 1
    local = breadth_counts[0] + breadth_counts[1]
    intra = breadth_counts[2]
    cross = breadth_counts[3] + breadth_counts[4]
    total_b = sum(breadth_counts.values()) or 1
    ws.cell(row_ptr, 2, "local (0-1)"); ws.cell(row_ptr, 3, local); ws.cell(row_ptr, 4, f"{local/total_b:.0%}")
    row_ptr += 1
    ws.cell(row_ptr, 2, "intra-contract (2)"); ws.cell(row_ptr, 3, intra); ws.cell(row_ptr, 4, f"{intra/total_b:.0%}")
    row_ptr += 1
    ws.cell(row_ptr, 2, "cross-boundary (3-4)"); ws.cell(row_ptr, 3, cross); ws.cell(row_ptr, 4, f"{cross/total_b:.0%}")
    row_ptr += 2

    # External spec Y/N
    ws.cell(row_ptr, 1, "External specification required -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)").font = bold
    row_ptr += 1
    ext_counts = {"Yes": 0, "No": 0}
    for row in rq2a_rows:
        v = row.get("external_spec")
        if v in ext_counts:
            ext_counts[v] += 1
    for k, v in ext_counts.items():
        ws.cell(row_ptr, 2, k); ws.cell(row_ptr, 3, v)
        row_ptr += 1
    row_ptr += 1

    # Additional functions distribution
    ws.cell(row_ptr, 1, "Additional functions required, distribution -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)").font = bold
    row_ptr += 1
    af_dist = {}
    for row in rq2a_rows:
        n = to_int(row.get("additional_functions_n"))
        if n is not None:
            af_dist[n] = af_dist.get(n, 0) + 1
    for k in sorted(af_dist):
        ws.cell(row_ptr, 2, f"{k} functions"); ws.cell(row_ptr, 3, af_dist[k])
        row_ptr += 1
    row_ptr += 1

    # Single vs multi
    ws.cell(row_ptr, 1, "Annotation multiplicity -- 46-case RQ2-A set (all Expressible=Yes, incl. not-yet-built)").font = bold
    row_ptr += 1
    mult_counts = {"single": 0, "multi": 0}
    for row in rq2a_rows:
        m = row.get("annotation_multiplicity")
        if m in mult_counts:
            mult_counts[m] += 1
    for k, v in mult_counts.items():
        ws.cell(row_ptr, 2, k); ws.cell(row_ptr, 3, v)
        row_ptr += 1
    row_ptr += 2

    # RQ1-C cross-tabs -- validation_outcome only exists for actually-run cases, so these use
    # validated_rows (45: baseline20 + case-built phase_reviews_expressible), not rq2a_rows (46).
    ws.cell(row_ptr, 1, "RQ1-C: value_or_algorithm x validation_outcome -- 45-case validated set").font = bold
    row_ptr += 1
    cross = {}
    for row in validated_rows:
        key = (row.get("value_or_algorithm") or "?", row.get("validation_outcome") or "?")
        cross[key] = cross.get(key, 0) + 1
    for (voa, outc), n in sorted(cross.items()):
        ws.cell(row_ptr, 2, voa); ws.cell(row_ptr, 3, outc); ws.cell(row_ptr, 4, n)
        row_ptr += 1
    row_ptr += 1

    ws.cell(row_ptr, 1, "RQ1-C: relation_form x validation_outcome -- 45-case validated set").font = bold
    row_ptr += 1
    cross2 = {}
    for row in validated_rows:
        key = (row.get("relation_form") or "?", row.get("validation_outcome") or "?")
        cross2[key] = cross2.get(key, 0) + 1
    for (rf, outc), n in sorted(cross2.items()):
        ws.cell(row_ptr, 2, rf); ws.cell(row_ptr, 3, outc); ws.cell(row_ptr, 4, n)
        row_ptr += 1
    row_ptr += 1

    ws.cell(row_ptr, 1, "RQ1-C: context_breadth x validation_outcome -- 45-case validated set").font = bold
    row_ptr += 1
    cross3 = {}
    for row in validated_rows:
        key = (to_int(row.get("context_breadth")), row.get("validation_outcome") or "?")
        cross3[key] = cross3.get(key, 0) + 1
    for (cb, outc), n in sorted(cross3.items(), key=lambda kv: (kv[0][0] is None, kv[0])):
        ws.cell(row_ptr, 2, cb); ws.cell(row_ptr, 3, outc); ws.cell(row_ptr, 4, n)
        row_ptr += 1
    row_ptr += 1

    if n_unbuilt:
        ws.cell(row_ptr, 1,
                f"Note: {n_unbuilt} Expressible=Yes case(s) have no built case JSON yet "
                f"(e.g. web3bugs_16_H_04) -- included in the 46-case RQ2-A structural stats above, "
                f"excluded from the 45-case validated/RQ1-C cross-tabs.")
        row_ptr += 1

    # Blocker tag distribution across all Inexpressible (34 cases: phase_reviews_inexpressible 9 + old_l4_track20)
    ws.cell(row_ptr, 1, "Blocker tag distribution -- all Inexpressible cases (phase_reviews_inexpressible + old_l4_track20)").font = bold
    row_ptr += 1
    inexpr_rows = [r for r in rows if r["group"] in ("phase_reviews_inexpressible", "old_l4_track20")]
    tag_dist = {}
    for row in inexpr_rows:
        t = row.get("blocker_tags") or "(none)"
        tag_dist[t] = tag_dist.get(t, 0) + 1
    for k in sorted(tag_dist):
        ws.cell(row_ptr, 2, k); ws.cell(row_ptr, 3, tag_dist[k])
        row_ptr += 1

    autosize(ws)
    return ws


def write_by_outcome(wb, rows):
    ws = wb.create_sheet("by_outcome")
    bold = Font(bold=True)
    # Only case-built, actually-validated cases have a real validation_outcome.
    validated_rows = [
        r for r in rows
        if r["group"] in ("baseline20", "phase_reviews_expressible") and r.get("case_built") == "Yes"
    ]

    buckets = {"Violated": [], "Warning": [], "Unsupported": [], "Satisfied": [], "Other": []}
    for row in validated_rows:
        o = row.get("validation_outcome") or "Other"
        buckets.setdefault(o, buckets.get(o, []))
        buckets[o if o in buckets else "Other"].append(row)

    ws.append(["outcome", "n_cases", "median_relevant_statements", "median_unique_values",
               "median_context_breadth", "external_spec_yes_pct"])
    for c in ws[1]:
        c.font = bold
    for outcome, grp in buckets.items():
        if not grp:
            continue
        rs = sorted(v for v in (to_int(r.get("relevant_statements")) for r in grp) if v is not None)
        uv = sorted(v for v in (to_int(r.get("unique_values")) for r in grp) if v is not None)
        cb = sorted(v for v in (to_int(r.get("context_breadth")) for r in grp) if v is not None)
        ext_yes = sum(1 for r in grp if r.get("external_spec") == "Yes")
        ws.append([
            outcome, len(grp),
            statistics.median(rs) if rs else "",
            statistics.median(uv) if uv else "",
            statistics.median(cb) if cb else "",
            f"{ext_yes/len(grp):.0%}" if grp else "",
        ])
    autosize(ws)
    return ws


def write_value_usable_matrix(wb, rows):
    ws = wb.create_sheet("value_usable_matrix")
    bold = Font(bold=True)
    eligible_rows = [r for r in rows if r["case_id"] != EXCLUDED_CASE_ID]

    ws.cell(1, 1, "Value/Algorithm x Usable/Unusable -- 74 eligible cases (all groups)").font = bold
    matrix = {("Value", "Usable"): 0, ("Value", "Unusable"): 0,
              ("Algorithm", "Usable"): 0, ("Algorithm", "Unusable"): 0}
    unclassified = 0
    for row in eligible_rows:
        key = (row.get("value_or_algorithm"), row.get("usable"))
        if key in matrix:
            matrix[key] += 1
        else:
            unclassified += 1

    r = 3
    ws.cell(r, 2, "Usable").font = bold
    ws.cell(r, 3, "Unusable").font = bold
    r += 1
    ws.cell(r, 1, "Value").font = bold
    ws.cell(r, 2, matrix[("Value", "Usable")])
    ws.cell(r, 3, matrix[("Value", "Unusable")])
    r += 1
    ws.cell(r, 1, "Algorithm").font = bold
    ws.cell(r, 2, matrix[("Algorithm", "Usable")])
    ws.cell(r, 3, matrix[("Algorithm", "Unusable")])
    r += 2
    if unclassified:
        ws.cell(r, 1, f"Unclassified/missing: {unclassified}")
        r += 2

    # breakdown by group
    ws.cell(r, 1, "Breakdown by group").font = bold
    r += 1
    ws.append_row = None  # no-op, keep openpyxl happy
    ws.cell(r, 1, "group"); ws.cell(r, 2, "Value/Usable"); ws.cell(r, 3, "Value/Unusable")
    ws.cell(r, 4, "Algorithm/Usable"); ws.cell(r, 5, "Algorithm/Unusable")
    for c in range(1, 6):
        ws.cell(r, c).font = bold
    r += 1
    groups = ["baseline20", "phase_reviews_expressible", "phase_reviews_inexpressible", "old_l4_track20"]
    for g in groups:
        grp_rows = [row for row in eligible_rows if row["group"] == g]
        m = {("Value", "Usable"): 0, ("Value", "Unusable"): 0,
             ("Algorithm", "Usable"): 0, ("Algorithm", "Unusable"): 0}
        for row in grp_rows:
            key = (row.get("value_or_algorithm"), row.get("usable"))
            if key in m:
                m[key] += 1
        ws.cell(r, 1, g)
        ws.cell(r, 2, m[("Value", "Usable")])
        ws.cell(r, 3, m[("Value", "Unusable")])
        ws.cell(r, 4, m[("Algorithm", "Usable")])
        ws.cell(r, 5, m[("Algorithm", "Unusable")])
        r += 1

    autosize(ws)
    return ws


def write_latency_sheet(wb):
    ws = wb.create_sheet("latency")
    if not LATENCY_CSV.exists():
        ws.cell(1, 1, "rq2b_latency.csv not found yet -- run collect_rq2b.py first")
        return ws
    with open(LATENCY_CSV, encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            ws.append(row)
            if i == 0:
                for c in ws[1]:
                    c.font = Font(bold=True)
    autosize(ws)

    ws2 = wb.create_sheet("latency_cumulative_median")
    if CUMMEDIAN_CSV.exists():
        with open(CUMMEDIAN_CSV, encoding="utf-8") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                ws2.append(row)
                if i == 0:
                    for c in ws2[1]:
                        c.font = Font(bold=True)
        autosize(ws2)
    return ws


def write_structural_sheet(wb):
    ws = wb.create_sheet("rq1_structural")
    if not STRUCTURAL_CSV.exists():
        ws.cell(1, 1, "rq1_structural.csv not found yet -- run collect_rq1_structural.py first")
        return ws
    with open(STRUCTURAL_CSV, encoding="utf-8") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            ws.append(row)
            if i == 0:
                for c in ws[1]:
                    c.font = Font(bold=True)
    autosize(ws)
    return ws


def main():
    extracted = load_extracted()
    latency = load_latency()
    cummedian = load_cummedian()
    validation = load_validation()

    print(f"[info] loaded {len(extracted)} extracted case records")
    print(f"[info] loaded {len(latency)} latency rows, {len(validation)} validation rows")

    expected_total = 74  # 20 baseline + 34 phase_reviews eligible + 20 old-L4
    if len(extracted) < expected_total:
        missing_hint = expected_total - len(extracted)
        print(f"[warn] only {len(extracted)}/{expected_total} eligible cases have an extracted "
              f"record ({missing_hint} missing) -- run with partial data anyway, re-run this "
              f"script once all extraction agents finish.")

    rows = merge_records(extracted, latency, validation)

    wb = Workbook()
    write_raw_profile(wb, rows)
    write_latency_sheet(wb)
    write_structural_sheet(wb)
    write_aggregates(wb, rows)
    write_by_outcome(wb, rows)
    write_value_usable_matrix(wb, rows)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"[ok] workbook written -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
